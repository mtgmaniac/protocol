#!/usr/bin/env python3
"""Export ability eff + inspect text; highlight rows changed since the prior review export."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT / "scripts" / "debug") not in sys.path:
    sys.path.insert(0, str(ROOT / "scripts" / "debug"))

from ability_text_format import format_eff, format_inspect  # noqa: E402

HEROES_PATH = ROOT / "data/raw/heroes.data.json"
ENEMIES_PATH = ROOT / "data/raw/enemies.data.json"
OUT_PATH = ROOT / "docs/ABILITY_TEXT_REVIEW.xlsx"
OUT_PATH_FALLBACK = ROOT / "docs/ABILITY_TEXT_REVIEW_v2.xlsx"

ENEMY_ZONE_RANGES = {
    "recharge": (1, 4),
    "strike": (5, 10),
    "surge": (11, 16),
    "crit": (17, 19),
    "overload": (20, 20),
}
ZONES = ["recharge", "strike", "surge", "crit", "overload"]

HEADERS = [
    "ID",
    "Side",
    "Unit ID",
    "Unit Name",
    "Form",
    "Zone",
    "Roll Min",
    "Roll Max",
    "Ability Name",
    "Current Eff",
    "Proposed Eff",
    "Current Inspect",
    "Proposed Inspect",
    "Changed This Pass",
    "Approved",
    "Notes",
]

ROW_KEY_FIELDS = ("Side", "Unit ID", "Form", "Zone", "Ability Name")


def row_key(row: dict[str, object]) -> tuple[str, ...]:
    return tuple(str(row.get(field, "")) for field in ROW_KEY_FIELDS)


def load_prior_proposed(path: Path) -> dict[tuple[str, ...], tuple[str, str]]:
    if not path.is_file():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    if "All Abilities" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["All Abilities"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h): i for i, h in enumerate(headers) if h}
    required = ["Side", "Unit ID", "Form", "Zone", "Ability Name", "Proposed Eff", "Proposed Inspect"]
    if not all(k in idx for k in required):
        wb.close()
        return {}
    prior: dict[tuple[str, ...], tuple[str, str]] = {}
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if not row_cells or row_cells[idx["Ability Name"]] is None:
            continue
        key_row = {
            "Side": row_cells[idx["Side"]],
            "Unit ID": row_cells[idx["Unit ID"]],
            "Form": row_cells[idx["Form"]],
            "Zone": row_cells[idx["Zone"]],
            "Ability Name": row_cells[idx["Ability Name"]],
        }
        key = row_key(key_row)
        prior[key] = (
            str(row_cells[idx["Proposed Eff"]] or ""),
            str(row_cells[idx["Proposed Inspect"]] or ""),
        )
    wb.close()
    return prior


def legacy_inspect(raw: dict, side: str) -> str:
    if not raw:
        return ""
    parts: list[str] = []
    dmg = int(raw.get("dmg", 0) or 0)
    d_min = int(raw.get("dMin", 0) or 0)
    d_max = int(raw.get("dMax", 0) or 0)
    if dmg > 0:
        suffix = " to all enemies" if raw.get("blastAll") else ""
        parts.append(f"Deal {dmg} damage{suffix}.")
    elif d_min > 0 or d_max > 0:
        parts.append(f"Deal {d_min}-{d_max} damage.")
    burn = int(raw.get("burn", 0) or 0)
    if burn > 0:
        dt = int(raw.get("burnT", 0) or 0)
        suffix = f" for {dt} turn{'s' if dt != 1 else ''}" if dt > 0 else ""
        parts.append(f"Deal {burn} damage per turn{suffix}.")
    heal = int(raw.get("heal", 0) or 0)
    if heal > 0:
        scope = " to all allies" if raw.get("healAll") else ""
        parts.append(f"Restore {heal} HP{scope}.")
    shield = int(raw.get("shield", 0) or 0)
    if shield > 0:
        parts.append(f"Gain {shield} shield this round.")
    rfe = int(raw.get("rfe", 0) or 0)
    if rfe > 0:
        parts.append(f"Reduce an enemy die by {rfe}.")
    rfm = int(raw.get("rfm", 0) or 0)
    if rfm > 0:
        parts.append(f"Raise a hero die by {rfm}.")
    if raw.get("revive") or raw.get("reviveAll"):
        parts.append(f"Revive a fallen ally at {int(raw.get('revivePct', 50))}% max HP.")
    if raw.get("cloak"):
        parts.append("Cloak: untargetable by hostile single-target abilities; breaks on dealing damage or an AoE hit.")
    if raw.get("taunt"):
        parts.append("Taunt: enemies must target this unit.")
    if raw.get("ignSh"):
        parts.append("Pierces enemy shields.")
    if not parts:
        return str(raw.get("eff", ""))
    return " ".join(parts)


def collect_rows(prior: dict[tuple[str, ...], tuple[str, str]]) -> list[dict[str, object]]:
    heroes = json.loads(HEROES_PATH.read_text(encoding="utf-8"))
    enemies = json.loads(ENEMIES_PATH.read_text(encoding="utf-8"))
    rows: list[dict[str, object]] = []

    def add_row(
        side: str,
        unit_id: str,
        unit_name: str,
        form: str,
        zone: str,
        roll_min: int,
        roll_max: int,
        name: str,
        raw: dict,
    ) -> None:
        current_eff = str(raw.get("eff", ""))
        proposed_eff = format_eff(raw, side)
        proposed_inspect = format_inspect(raw, side)
        current_inspect = legacy_inspect(raw, side)
        row: dict[str, object] = {
            "Side": side,
            "Unit ID": unit_id,
            "Unit Name": unit_name,
            "Form": form,
            "Zone": zone,
            "Roll Min": roll_min,
            "Roll Max": roll_max,
            "Ability Name": name,
            "Current Eff": current_eff,
            "Proposed Eff": proposed_eff,
            "Current Inspect": current_inspect,
            "Proposed Inspect": proposed_inspect,
            "Approved": "",
            "Notes": "",
        }
        key = row_key(row)
        old_eff, old_inspect = prior.get(key, ("", ""))
        changed = (
            proposed_eff != old_eff
            or proposed_inspect != old_inspect
            or (not prior and (current_eff != proposed_eff or current_inspect != proposed_inspect))
        )
        row["Changed This Pass"] = "Y" if changed else "N"
        rows.append(row)

    for hero in heroes["heroes"]:
        hid = str(hero["id"])
        hname = str(hero["name"])
        for ab in hero.get("abilities", []):
            rng = ab.get("range", [0, 0])
            add_row(
                "hero",
                hid,
                hname,
                "Base",
                str(ab.get("zone", "")),
                int(rng[0]) if len(rng) > 0 else 0,
                int(rng[1]) if len(rng) > 1 else 0,
                str(ab.get("name", "")),
                ab,
            )
        for evo in hero.get("evolutions", []):
            ename = str(evo.get("name", evo.get("id", "Evolution")))
            for ab in evo.get("abilities", []):
                rng = ab.get("range", [0, 0])
                add_row(
                    "hero",
                    hid,
                    hname,
                    ename,
                    str(ab.get("zone", "")),
                    int(rng[0]) if len(rng) > 0 else 0,
                    int(rng[1]) if len(rng) > 1 else 0,
                    str(ab.get("name", "")),
                    ab,
                )

    for etype, suite in enemies["enemyAbilities"].items():
        for zone in ZONES:
            ab = suite.get(zone)
            if not ab:
                continue
            lo, hi = ENEMY_ZONE_RANGES[zone]
            add_row(
                "enemy",
                str(etype),
                str(etype),
                "Type suite",
                zone,
                lo,
                hi,
                str(ab.get("name", "")),
                ab,
            )

    for i, row in enumerate(rows, start=1):
        row["ID"] = i
    return rows


def build_workbook(rows: list[dict[str, object]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "All Abilities"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    changed_fill = PatternFill("solid", fgColor="FEF3C7")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=2):
        highlight = row.get("Changed This Pass") == "Y"
        for c_idx, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if highlight:
                cell.fill = changed_fill

    widths = {
        "A": 6,
        "B": 8,
        "C": 14,
        "D": 18,
        "E": 20,
        "F": 12,
        "G": 9,
        "H": 9,
        "I": 24,
        "J": 42,
        "K": 42,
        "L": 48,
        "M": 56,
        "N": 16,
        "O": 10,
        "P": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    summary = wb.create_sheet("Summary")
    changed = sum(1 for r in rows if r["Changed This Pass"] == "Y")
    summary.append(["Metric", "Count"])
    summary.append(["Total abilities", len(rows)])
    summary.append(["Highlighted (changed this pass)", changed])
    summary.append(["", ""])
    summary.append(["Highlight legend", ""])
    summary.append(["Yellow row", "Proposed eff/inspect differs from prior export"])
    summary.append(["Data fixes applied", "Mandible Rake 8 dmg; Ramming Plate 12 dmg; Culling Feed no heal"])
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 12

    return wb


def main() -> None:
    prior_path = OUT_PATH if OUT_PATH.is_file() else OUT_PATH_FALLBACK
    prior = load_prior_proposed(prior_path)
    rows = collect_rows(prior)
    wb = build_workbook(rows)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    out = OUT_PATH
    try:
        wb.save(out)
    except PermissionError:
        out = OUT_PATH_FALLBACK
        wb.save(out)
    changed = sum(1 for r in rows if r["Changed This Pass"] == "Y")
    print(f"Wrote {out}")
    print(f"Rows: {len(rows)}, highlighted this pass: {changed}")


if __name__ == "__main__":
    main()
